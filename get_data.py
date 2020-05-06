'''
Name: Automation data collection

What the programme does: Goes to NEA website
                         Check if the collection from the website is 5pm
                         If it is, collect data for SO2
                         input into excel database
                         change data to PM10
                         collect and input data into excel database in different sheet
                         close programme
                         
                         if the time is before 5pm,
                         the programme  will wait for 5 min before trying the programme again
                         and it will repeat until the time the data start is at 5pm
                        
                         if the time is after 5pm,
                         the programme will ask the user to collect the data manually
'''
from selenium import webdriver
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import time

mypath = os.path.dirname(os.path.abspath(__file__))+'\\'

def copydata_pollutant(row,name,pollutant):  
    region=[]
    time1=[]
    time2=[]
    time3=[]
    time4=[]
    time5=[]
    time6=[]
    time7=[]
    time8=[]
    time9=[]
    time10=[]
    time11=[]
    time12=[]
    if name == 'firsthalf':
        db_df = pd.read_excel('database.xlsx',pollutant)
        row_excel = len(db_df)+4
        wb=openpyxl.load_workbook('database.xlsx')
        sheet=wb.get_sheet_by_name(pollutant)
        today = datetime.today().strftime('%d/%m/%Y')
        sheet.cell(row=row_excel-1,column = 1).value=today
        wb.save('database.xlsx')

        
        for r in range(2,row+1):
            for c in range(1,14):
                value = driver.find_element_by_xpath('//*[@id="pollutantConDetail-1"]/tbody/tr['+str(r)+']/td['+str(c)+']').text
                if c == 1:
                    region.append(value)
                elif c ==2:
                    time1.append(value)
                elif c ==3:
                    time2.append(value)
                elif c ==4:
                    time3.append(value)
                elif c ==5:
                    time4.append(value)
                elif c ==6:
                    time5.append(value)
                elif c ==7:
                    time6.append(value)
                elif c ==8:
                    time7.append(value)    
                elif c ==9:
                    time8.append(value)                
                elif c ==10:
                    time9.append(value)
                elif c ==11:
                    time10.append(value)
                elif c ==12:
                    time11.append(value)
                elif c ==13:
                    time12.append(value)
    
        data = {'Region/Time': region, '6pm':time1,'7pm':time2,'8pm':time3,'9pm':time4,'10pm':time5,
                '11pm':time6,'12am':time7,'1am':time8,'2am':time9,'3am':time10,'4am':time11,'5am':time12}
        data_df = pd.DataFrame.from_dict(data)
        
        book = load_workbook('database.xlsx')
        writer = pd.ExcelWriter('database.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        data_df.to_excel(writer,sheet_name=pollutant, index=False, startrow =writer.sheets[pollutant].max_row)
        writer.save()
        
        
    if name == 'secondhalf':
        for r in range(2,row+1):
            for c in range(1,14):
                value = driver.find_element_by_xpath('//*[@id="pollutantConDetail-2"]/tbody/tr['+str(r)+']/td['+str(c)+']').text
                if c == 1:
                    region.append(value)
                elif c ==2:
                    time1.append(value)
                elif c ==3:
                    time2.append(value)
                elif c ==4:
                    time3.append(value)
                elif c ==5:
                    time4.append(value)
                elif c ==6:
                    time5.append(value)
                elif c ==7:
                    time6.append(value)
                elif c ==8:
                    time7.append(value)    
                elif c ==9:
                    time8.append(value)                
                elif c ==10:
                    time9.append(value)
                elif c ==11:
                    time10.append(value)
                elif c ==12:
                    time11.append(value)
                elif c ==13:
                    time12.append(value)
    
        data = {'Region/Time': region, '6am':time1,'7am':time2,'8am':time3,'9am':time4,'10am':time5,
                '11am':time6,'12am':time7,'1pm':time8,'2pm':time9,'3pm':time10,'4pm':time11,'5pm':time12}
        data_df = pd.DataFrame.from_dict(data)
        book = load_workbook('database.xlsx')
        writer = pd.ExcelWriter('database.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        data_df.to_excel(writer,sheet_name=pollutant, index=False, startrow =writer.sheets[pollutant].max_row+1)
        writer.save()
        
def copydata_5pm(pollutant='overall'):
    
    db_df = pd.read_excel('database.xlsx',pollutant)
    row_excel = len(db_df)+4
    wb=openpyxl.load_workbook('database.xlsx')
    sheet=wb.get_sheet_by_name(pollutant)
    today = datetime.today().strftime('%d/%m/%Y')
    sheet.cell(row=row_excel-1,column = 1).value=today
    wb.save('database.xlsx')
    # getting readings on 5pm 
    row = len(driver.find_elements_by_xpath('//*[@id="pollutantConSummary"]/tbody/tr'))
    region=[]
    so=[]
    pm=[]
    
    for r in range(2,row+1):
        for c in range(1,4):
            value = driver.find_element_by_xpath('//*[@id="pollutantConSummary"]/tbody/tr['+str(r)+']/td['+str(c)+']').text
            if c == 1:
                region.append(value)
            elif c ==2:
                so.append(value)
            else:
                pm.append(value)
    
    
    #df_bs = pd.DataFrame(data,columns=['Regions','24-hr Sulphur Dioxide (µg/m3)','24-hr PM10 (µg/m3)'])
    data = {'Region/Time': region, '24-hr Sulphur Dioxide (µg/m3)': so,'24-hr PM10 (µg/m3)': pm}
    data_df = pd.DataFrame.from_dict(data)
    book = load_workbook('database.xlsx')
    writer = pd.ExcelWriter('database.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    data_df.to_excel(writer,sheet_name=pollutant, index=False, startrow =writer.sheets[pollutant].max_row)
    writer.save()
    
    
#start of programme    
if __name__ == "__main__":
    loop = 0

    while loop == 0:     
        # connect to windows or mac chromedriver
        if os.name == 'nt':
            cdriver = mypath+'chromedriver.exe'
        else: # for macOS
            cdriver = mypath+'chromedriver_mac'
        
        
        url='https://www.haze.gov.sg/resources/pollutant-concentrations'
        driver=webdriver.Chrome(cdriver)
        driver.get(url) 
        
        # checking if the data starts at 5pm:
        check=driver.find_element_by_xpath('//*[@id="pollutantConDetail-1"]/tbody/tr[1]/th[2]').text
        if check == '6pm':
            print('Starting collecting data ...')
            
            row = len(driver.find_elements_by_xpath('//*[@id="pollutantConDetail-1"]/tbody/tr'))
            print('Collection of SO2 data in process...' )
            copydata_pollutant(row,name = 'firsthalf',pollutant = 'SO2')
            copydata_pollutant(row,name = 'secondhalf',pollutant = 'SO2')
            print('\nCollection of SO2 data completed' )
            print('\nCollection of PM10 data in process...' )
            el = driver.find_element_by_id('typeReading')
            for option in el.find_elements_by_tag_name('option'):
                if option.text == '24-hr PM10':
                    option.click() 
                    
            
            copydata_pollutant(row,name = 'firsthalf',pollutant = 'PM10')
            copydata_pollutant(row,name = 'secondhalf',pollutant = 'PM10')
            print('\nCollection of PM10 data completed' )
            
            print('\nGetting 5pm data')
            copydata_5pm()
            print('\nData extractiong from NEA website has been completed.')
            
            driver.quit()
            break
    
        elif check =='7pm':
            print('The data starts at '+check+', it is not 6pm')
            print('You have missed 1 hour data. Please collect the data yourself')
            driver.quit()
            break
        
        else:
            print('the data starts at '+check+', it is not 6pm')
            print('System will try retrive result again in 5 mins')
            driver.quit()
            time.sleep(300)
